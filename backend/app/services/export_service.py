from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.campsite import Campsite
from app.models.trip import Trip

EXPORT_FORMAT_VERSION = "1.2"
HEADER_FONT = Font(bold=True)

TRIPS_HEADERS = [
    "trip_id",
    "location_name",
    "place_area",
    "plot_number",
    "country",
    "stay_type",
    "latitude",
    "longitude",
    "start_date",
    "end_date",
    "nights",
    "price_total_sek",
    "price_per_night_input_sek",
    "price_input_mode",
    "currency",
    "star_rating",
    "notes",
    "campsite_id",
    "photo_count",
    "photo_filenames",
    "created_at",
    "updated_at",
]

CAMPSITE_HEADERS = ["campsite_id", "name", "latitude", "longitude", "notes", "created_at"]

README_ROWS = [
    ("Trips", "trip_id", "text (uuid)", "no", "Unique identifier for the trip"),
    ("Trips", "location_name", "text", "no", "Name of the campsite / location"),
    ("Trips", "place_area", "text", "yes", "Broader area/region, if recorded"),
    ("Trips", "plot_number", "text", "yes", "Plot/site identifier as free text (not always numeric)"),
    ("Trips", "country", "text", "yes", "Country the trip took place in, free text"),
    ("Trips", "stay_type", "text", "yes", "One of: camping, stallplats, fricamping"),
    ("Trips", "latitude", "number", "yes", "Decimal degrees, WGS84"),
    ("Trips", "longitude", "number", "yes", "Decimal degrees, WGS84"),
    ("Trips", "start_date", "date", "no", "First night of the stay"),
    ("Trips", "end_date", "date", "no", "Checkout date (exclusive) - nights = end_date - start_date"),
    ("Trips", "nights", "integer", "no", "Derived: end_date - start_date. Never stored directly."),
    ("Trips", "price_total_sek", "number", "yes", "Total price for the whole stay"),
    (
        "Trips",
        "price_per_night_input_sek",
        "number",
        "yes",
        "Price per night as originally entered, if that's how it was recorded",
    ),
    ("Trips", "price_input_mode", "text", "no", "How price was entered: total, per_night, or none"),
    ("Trips", "currency", "text", "no", "Currency code, e.g. SEK"),
    ("Trips", "star_rating", "integer", "yes", "1-5 stars"),
    ("Trips", "notes", "text", "yes", "Free-text notes"),
    ("Trips", "campsite_id", "text (uuid)", "yes", "Link to the Campsites sheet, if this trip was matched to a reusable place"),
    (
        "Trips",
        "photo_count",
        "integer",
        "no",
        "Number of photos attached to this trip. Photo files themselves are NOT included in this export.",
    ),
    ("Trips", "photo_filenames", "text", "yes", "Comma-separated filenames of attached photos, for reference only"),
    ("Trips", "created_at", "datetime (UTC)", "no", "When the trip record was first created"),
    ("Trips", "updated_at", "datetime (UTC)", "no", "When the trip record was last edited"),
    ("Campsites", "campsite_id", "text (uuid)", "no", "Unique identifier for the campsite"),
    ("Campsites", "name", "text", "no", "Campsite name"),
    ("Campsites", "latitude", "number", "yes", "Decimal degrees, WGS84"),
    ("Campsites", "longitude", "number", "yes", "Decimal degrees, WGS84"),
    ("Campsites", "notes", "text", "yes", "Free-text notes"),
    ("Campsites", "created_at", "datetime (UTC)", "no", "When the campsite record was created"),
]


def _naive(dt: datetime | None) -> datetime | None:
    return dt.replace(tzinfo=None) if dt is not None else None


def _trips_in_range(db: Session, start_date: date | None, end_date: date | None) -> list[Trip]:
    stmt = select(Trip).options(selectinload(Trip.photos)).order_by(Trip.start_date)
    if start_date is not None:
        stmt = stmt.where(Trip.end_date > start_date)
    if end_date is not None:
        stmt = stmt.where(Trip.start_date < end_date)
    return list(db.scalars(stmt).all())


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = HEADER_FONT
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"


def _autosize_columns(ws: Worksheet) -> None:
    for column_cells in ws.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)


def _build_trips_sheet(ws: Worksheet, trips: list[Trip]) -> None:
    _write_header(ws, TRIPS_HEADERS)
    for trip in trips:
        ws.append(
            [
                str(trip.id),
                trip.location_name,
                trip.place_area,
                trip.plot_number,
                trip.country,
                trip.stay_type.value if trip.stay_type is not None else None,
                trip.latitude,
                trip.longitude,
                trip.start_date,
                trip.end_date,
                trip.nights,
                float(trip.price_total) if trip.price_total is not None else None,
                float(trip.price_per_night_input) if trip.price_per_night_input is not None else None,
                trip.price_input_mode.value,
                trip.currency,
                trip.star_rating,
                trip.notes,
                str(trip.campsite_id) if trip.campsite_id else None,
                len(trip.photos),
                ", ".join(photo.original_filename or photo.file_path for photo in trip.photos) or None,
                _naive(trip.created_at),
                _naive(trip.updated_at),
            ]
        )
    _autosize_columns(ws)


def _build_campsites_sheet(ws: Worksheet, campsites: list[Campsite]) -> None:
    _write_header(ws, CAMPSITE_HEADERS)
    for campsite in campsites:
        ws.append(
            [
                str(campsite.id),
                campsite.name,
                campsite.latitude,
                campsite.longitude,
                campsite.notes,
                _naive(campsite.created_at),
            ]
        )
    _autosize_columns(ws)


def _build_readme_sheet(ws: Worksheet, include_campsites: bool, generated_at: datetime, trip_count: int) -> None:
    ws.append([f"Camping Logbook export - format version {EXPORT_FORMAT_VERSION}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Generated at {generated_at.isoformat()}"])
    ws.append([f"Trips included: {trip_count}"])
    ws.append([])
    ws.append(["Photos are not included in this export.", "See photo_count / photo_filenames on the Trips sheet."])
    ws.append(["Nights are derived, not stored.", "nights = end_date - start_date."])
    ws.append([])

    _write_header(ws, ["sheet", "column", "type", "nullable", "description"])
    for row in README_ROWS:
        if row[0] == "Campsites" and not include_campsites:
            continue
        ws.append(row)
    _autosize_columns(ws)


def generate_export_workbook(db: Session, start_date: date | None, end_date: date | None) -> BytesIO:
    trips = _trips_in_range(db, start_date, end_date)
    campsites = list(db.scalars(select(Campsite).order_by(Campsite.name)).all())
    generated_at = datetime.now(timezone.utc)

    workbook = Workbook()
    trips_sheet = workbook.active
    trips_sheet.title = "Trips"
    _build_trips_sheet(trips_sheet, trips)

    if campsites:
        _build_campsites_sheet(workbook.create_sheet("Campsites"), campsites)

    _build_readme_sheet(workbook.create_sheet("Schema_ReadMe"), bool(campsites), generated_at, len(trips))

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
