import base64
import io

from openpyxl import load_workbook

from app.models.campsite import Campsite

TINY_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def _create_trip(client, auth_headers, **overrides):
    payload = {
        "location_name": "Export Camp",
        "start_date": "2025-06-01",
        "end_date": "2025-06-02",
        "price_input_mode": "none",
    }
    payload.update(overrides)
    response = client.post("/api/trips", json=payload, headers=auth_headers)
    assert response.status_code == 201, response.text
    return response.json()


def _load(response):
    return load_workbook(io.BytesIO(response.content))


def test_export_requires_auth(client):
    response = client.get("/api/export/xlsx")
    assert response.status_code == 401


def test_export_headers(client, auth_headers):
    response = client.get("/api/export/xlsx", headers=auth_headers)
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment; filename=" in response.headers["content-disposition"]


def test_export_trips_sheet_contents(client, auth_headers):
    trip = _create_trip(
        client,
        auth_headers,
        location_name="Trosa Havsbad och Camping",
        plot_number="420",
        country="Sweden",
        stay_type="camping",
        start_date="2025-08-06",
        end_date="2025-08-08",
        price_per_night_input=495,
        price_input_mode="per_night",
        star_rating=4,
        notes="Fint lage",
    )

    response = client.get("/api/export/xlsx", headers=auth_headers)
    wb = _load(response)

    assert "Trips" in wb.sheetnames
    ws = wb["Trips"]
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "trip_id",
        "location_name",
        "place_area",
        "plot_number",
        "country",
        "stay_type",
        "latitude",
        "longitude",
        "start_date",
        "end_date",
        "nights",
        "price_total_sek",
        "price_per_night_input_sek",
        "price_input_mode",
        "currency",
        "star_rating",
        "notes",
        "campsite_id",
        "photo_count",
        "photo_filenames",
        "created_at",
        "updated_at",
    ]

    row2 = tuple(cell.value for cell in ws[2])
    assert row2[0] is not None, "row 2 must be a real data row, not a blank row after the header"

    # Search by id rather than assuming position: other trips may already exist
    # (this runs against a shared dev database, not a pristine one) and sort
    # before this one by start_date.
    row = next(r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == trip["id"])
    by_header = dict(zip(headers, row))
    assert by_header["location_name"] == "Trosa Havsbad och Camping"
    assert by_header["plot_number"] == "420"
    assert by_header["country"] == "Sweden"
    assert by_header["stay_type"] == "camping"
    assert by_header["nights"] == 2
    assert by_header["price_total_sek"] == 990.0
    assert by_header["price_per_night_input_sek"] == 495.0
    assert by_header["star_rating"] == 4
    assert by_header["photo_count"] == 0
    assert by_header["photo_filenames"] is None
    assert ws.auto_filter.ref is not None
    assert ws.freeze_panes == "A2"


def test_export_includes_photo_metadata_not_photo_bytes(client, auth_headers):
    trip = _create_trip(client, auth_headers, location_name="Photo Export Camp")
    client.post(
        f"/api/trips/{trip['id']}/photos",
        files={"file": ("beach.png", io.BytesIO(TINY_PNG), "image/png")},
        headers=auth_headers,
    )

    response = client.get("/api/export/xlsx", headers=auth_headers)
    wb = _load(response)
    ws = wb["Trips"]
    headers = [cell.value for cell in ws[1]]
    row = next(r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == trip["id"])
    by_header = dict(zip(headers, row))
    assert by_header["photo_count"] == 1
    assert by_header["photo_filenames"] == "beach.png"


def test_export_date_range_filters_trips(client, auth_headers):
    _create_trip(client, auth_headers, location_name="In Range", start_date="2025-01-01", end_date="2025-01-02")
    _create_trip(client, auth_headers, location_name="Out Of Range", start_date="1999-01-01", end_date="1999-01-02")

    response = client.get(
        "/api/export/xlsx",
        params={"start_date": "2020-01-01", "end_date": "2030-01-01"},
        headers=auth_headers,
    )
    wb = _load(response)
    ws = wb["Trips"]
    names = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "In Range" in names
    assert "Out Of Range" not in names


def test_export_all_time_default_includes_everything(client, auth_headers):
    _create_trip(client, auth_headers, location_name="Ancient Trip", start_date="1990-01-01", end_date="1990-01-02")

    response = client.get("/api/export/xlsx", headers=auth_headers)
    wb = _load(response)
    ws = wb["Trips"]
    names = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "Ancient Trip" in names


def test_export_rejects_invalid_range(client, auth_headers):
    response = client.get(
        "/api/export/xlsx",
        params={"start_date": "2023-01-01", "end_date": "2023-01-01"},
        headers=auth_headers,
    )
    assert response.status_code == 422


def test_export_omits_campsites_sheet_when_none_exist(client, auth_headers):
    response = client.get("/api/export/xlsx", headers=auth_headers)
    wb = _load(response)
    assert "Campsites" not in wb.sheetnames
    assert "Schema_ReadMe" in wb.sheetnames
    readme_sheet = wb["Schema_ReadMe"]
    readme_rows = [tuple(r) for r in readme_sheet.iter_rows(values_only=True)]
    assert not any(r and r[0] == "Campsites" for r in readme_rows)


def test_export_includes_campsites_sheet_when_present(client, auth_headers, db_session):
    campsite = Campsite(name="Trosa Havsbad", latitude=58.9, longitude=17.5)
    db_session.add(campsite)
    db_session.commit()

    response = client.get("/api/export/xlsx", headers=auth_headers)
    wb = _load(response)
    assert "Campsites" in wb.sheetnames
    ws = wb["Campsites"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["campsite_id", "name", "latitude", "longitude", "notes", "created_at"]
    row = next(r for r in ws.iter_rows(min_row=2, values_only=True))
    assert row[1] == "Trosa Havsbad"

    readme_rows = [tuple(r) for r in wb["Schema_ReadMe"].iter_rows(values_only=True)]
    assert any(r and r[0] == "Campsites" for r in readme_rows)


def test_export_readme_sheet_mentions_photo_exclusion_and_nights_rule(client, auth_headers):
    response = client.get("/api/export/xlsx", headers=auth_headers)
    wb = _load(response)
    readme_text = " ".join(
        str(cell) for row in wb["Schema_ReadMe"].iter_rows(values_only=True) for cell in row if cell is not None
    )
    assert "not included in this export" in readme_text
    assert "nights = end_date - start_date" in readme_text
